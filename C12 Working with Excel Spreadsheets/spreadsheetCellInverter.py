#! python3
# program to invert the row and column of the cells in the spreadsheet

import openpyxl, sys

# check if the number of arguments is valid
if(len(sys.argv) != 2):
    print("Invalid number of arguments")
    sys.exit(1)

# load workbook
print("Loading workbook... " + str(sys.argv[1]))
try:
    wb = openpyxl.load_workbook(sys.argv[1])
except FileNotFoundError as err:
    print(str(err))
    sys.exit(1)
sheet = wb.active

# create data structure for spreadsheet data
wbLst = []

# loop through the spreadsheet and store the data in list
# every list in this list is a column from the sheet
print("Reading data from a file...")
for icol in range(1, sheet.max_column + 1):
    colLst = []
    for irow in range(1, sheet.max_row + 1):
        colLst.append(sheet.cell(row = irow, column = icol).value)
    wbLst.append(colLst)

# create new workbook
nwb = openpyxl.Workbook()
ns = nwb.active
ns.title = sheet.title

# write the data from wbLst into new worksheet
# we loop through each column and value in this column and write it row by row
# left to right
print("Writing data to new file...")
for x in range(0, len(wbLst)):
    for y in range(0, len(wbLst[x])):
        ns.cell(row = x+1, column = y+1).value = wbLst[x][y]

print("Saving file...")
try:
    nwb.save('inverted.xlsx')
except ValueError as err:
    print(str(err) + " Max number of columns is 18279")
    sys.exit(1)
print("Done")