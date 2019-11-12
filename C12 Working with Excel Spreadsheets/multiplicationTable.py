#! python3
# Create a program multiplicationTable.py that takes a number N from the command
# line and creates an N×N multiplication table in an Excel spreadsheet.

''' Check command line argument validity
    Create a workbook
    Create a multiplication table
        Print first row and column values with bold attribute
        Print other rows and columns by multiplicating each column by row value
'''
import openpyxl, sys
from openpyxl.styles import Font

# Check command line argument validity
# check number of arguments
if(len(sys.argv) != 2):
    print("Invalid number of arguments")
    sys.exit(1)
nArg = sys.argv[1]
# check if argument can be parsed as an integer
try:
    n = int(nArg)
except ValueError as err:
    print("Unable to parse argument as an integer: " + str(err))
    sys.exit(1)

# Create a workbook
wb = openpyxl.Workbook()
ws = wb.active
fontVar = Font(bold=True)

# Create a multiplication table
# Print first row and column values with bold attribute
for i in range(1, n + 1):
    cr = ws.cell(row = 1, column = i + 1)
    cc = ws.cell(row = i + 1, column = 1)
    cr.font = fontVar    # bold
    cc.font = fontVar    # bold
    cr.value = i
    cc.value = i

# Print other rows and columns by multiplicating each column by row value
for rowi in range(2, n + 2):
    for col in range(2, n + 2):
        ws.cell(row = rowi, column = col).value = (rowi - 1) * (col - 1)

wb.save('./multi.xlsx')
