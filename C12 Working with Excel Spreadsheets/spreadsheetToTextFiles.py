#! python3
# reads data from a spreadsheet and stores it in a text file
# this version assumes that filenames are stored in first row

# 1. Validate the command line argument's number
# 2. Check if given spreadsheet exists and terminate program if not
# 3. Set up the spreadsheet
# 4. Read spreadsheet column by column
#       open new text file with title from row 1
#       write each row from this column to this text file
#       close the file

import openpyxl, sys
from openpyxl.utils.cell import get_column_letter

# Validate the command line argument's number
if(len(sys.argv) != 2):
    print("Invalid number of arguments")
    sys.exit(1)

# Check if given spreadsheet exists and terminate program if not
try:
    wb = openpyxl.load_workbook(sys.argv[1])
except FileNotFoundError as err:
    print(str(err))
    sys.exit(1)

# Set up the spreadsheet
sheet = wb.active

# Read spreadsheet column by column
for icol in range(sheet.max_column):
    # open new text file with title from row 1
    title = sheet.cell(row=1, column=icol + 1).value
    print("Processing file " + title + "...")
    fileHandle = open(title, 'w')
    # write each row from this column to this text file
    for row in sheet[get_column_letter(icol+1)]:
        if(row.value == None or row.row == 1):
            continue
        fileHandle.write(str(row.value) + "\n")
    # close the file
    fileHandle.close()

print("Done")
