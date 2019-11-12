#! python3
# reads data from text files and writes it to a spreadsheet, each
# line is written to a separate row, and each file to separate column

# 1. Make a list of files from current working directory
# 2. Pick .txt files from the list
#       if there are no text files prompt message and terminate the program
# 3. Create a workbook and set it up
# 4. Loop through each file
#       open the file
#       read it line by line and store each line in corresponding row and col
# 5. Save the workbook to a file

import openpyxl
import os.path
import sys

# Make a list of files from current working directory
fileList = os.listdir('./')

# Pick .txt files from the list
textFList = []
for file in fileList:
    if(file.endswith('.txt')):
        textFList.append(file)

# if there are no text files prompt message and terminate the program
if not len(textFList):
    print("There are no text files in this directory")
    sys.exit(1)

# Create a workbook and set it up
wb = openpyxl.Workbook()
sheet = wb.active

# Loop through each file
for file in range(len(textFList)):
    print("Processing file: " + textFList[file] + "...")
    # open the file
    fileHandle = open(textFList[file], 'r')

    # read it line by line and store each line in corresponding row and col
    linesList = fileHandle.readlines()
    sheet.cell(row=1, column=file + 1).value = textFList[file].rstrip()
    for line in range(len(linesList)):
        sheet.cell(row=line + 2, column=file +
                   1).value = linesList[line].rstrip()

    fileHandle.close()
    
# Save the workbook to a file
print("Saving data...")
wb.save('textSpreadsheet.xlsx')
print("Done")
